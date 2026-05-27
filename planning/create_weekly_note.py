from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "週刊かつお"

# Colors
header_fill = PatternFill("solid", start_color="1F3864")
done_fill = PatternFill("solid", start_color="D9EAD3")
next_fill = PatternFill("solid", start_color="FCE5CD")
planned_fill = PatternFill("solid", start_color="EAF1FB")

thin = Side(style='thin', color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Headers
headers = ["#", "公開予定日", "タイトル案", "概要・メモ", "ステータス", "公開URL"]
col_widths = [5, 14, 30, 60, 12, 35]

for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
    ws.column_dimensions[get_column_letter(col)].width = w

ws.row_dimensions[1].height = 30

# Data
rows = [
    (1, "2026-03-21頃", "カナエルアプリを作った話",
     "自分でiOSアプリ「カナエル」を開発した経緯と感想。夢を叶えるためのアプリ。",
     "公開済み", ""),
    (2, "2026-03-28頃", "スペインに移住する話",
     "スペイン移住を考えている背景、理由、準備状況など。",
     "次回", ""),
    (3, "2026-04-04頃", "HikaruがイギリスのBoarding受験を決めた話",
     "息子HikaruがUKのボーディングスクール受験を決意。その経緯や親としての思い。",
     "予定", ""),
    (4, "2026-04-11頃", "iPadを制限したら娘がクリエイティブになった話",
     "キッズiPadの時間制限を設けたところ、漫画を描いたり読書をするようになった。\n"
     "デジタルが奪っていたのは時間でも思考力でもなく「クリエイティブ力」ではないか？",
     "予定", ""),
]

status_fills = {
    "公開済み": done_fill,
    "次回": next_fill,
    "予定": planned_fill,
}

for r_idx, row in enumerate(rows, 2):
    status = row[4]
    row_fill = status_fills.get(status, planned_fill)

    for c_idx, val in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
        if c_idx != 5:
            cell.fill = row_fill

    # Status cell with distinct color
    status_cell = ws.cell(row=r_idx, column=5)
    status_cell.fill = status_fills.get(status, planned_fill)
    status_cell.alignment = Alignment(horizontal="center", vertical="top")
    status_cell.font = Font(name="Arial", size=10, bold=True)

    ws.row_dimensions[r_idx].height = 60

# Freeze header row
ws.freeze_panes = "A2"

# Legend sheet
ws2 = wb.create_sheet("凡例")
ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 40

legend_header = ws2.cell(row=1, column=1, value="ステータス凡例")
legend_header.font = Font(name="Arial", bold=True, size=11)

legend_items = [
    ("公開済み", "すでにnoteに公開した記事"),
    ("次回", "次に公開予定（執筆対象）"),
    ("予定", "今後書く予定のネタ"),
]
fills = [done_fill, next_fill, planned_fill]

for i, ((label, desc), fill) in enumerate(zip(legend_items, fills), 2):
    c1 = ws2.cell(row=i, column=1, value=label)
    c1.fill = fill
    c1.font = Font(name="Arial", size=10, bold=True)
    c1.border = border
    c2 = ws2.cell(row=i, column=2, value=desc)
    c2.font = Font(name="Arial", size=10)
    c2.border = border

wb.save("/Users/katsuo/my-projects/sns-strategy/planning/週刊かつお_ネタ帳.xlsx")
print("Done")
