from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = load_workbook("/Users/katsuo/my-projects/sns-strategy/planning/週刊かつお_ネタ帳.xlsx")
ws = wb["週刊かつお"]

thin = Side(style='thin', color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
planned_fill = PatternFill("solid", start_color="EAF1FB")

new_topics = [
    # (タイトル案, カテゴリ, 概要)
    ("バツ1になった日の話", "バツ2・恋愛系", "初めての離婚。あの日何を感じたか、何を思ったか。キャラの根幹となるエピソード。"),
    ("ヒモをやっていた頃の本音", "バツ2・恋愛系", "ヒモ前科の実態。恥ずかしがらず、キャラとして使い倒す。"),
    ("麻美と再婚を決めた理由", "バツ2・恋愛系", "バツ2が再婚を決意した経緯。麻美への本音の愛情。"),
    ("無一文になったとき何を考えていたか", "無一文→FIRE", "どん底のリアル。そこから何を見ていたか。"),
    ("会社を売った日の翌朝", "無一文→FIRE", "13億売り上げた男が「どうしよう」ってなる瞬間。スタイルガイドに例文あり。"),
    ("FIREして「自由」になったのに何もできなかった話", "無一文→FIRE", "FIREの理想と現実のギャップ。自由の重さ。"),
    ("世界一周中に夢について考えたこと", "世界一周", "旅の中で夢とは何かを考えた体験談。カナエルの原点にもなりうる。"),
    ("「普通の国」より「変な国」の方が学びが多い話", "世界一周", "世界一周の体験から得た価値観。変わった国のエピソード。"),
    ("スペイン移住を子どもたちに伝えた日", "子育て・家族", "子どもたちの反応。家族の会話。"),
    ("Hikaruに夢を語ったら逆に論波された話", "子育て・家族", "息子Hikaruとの会話。子どもの視点に気づかされた話。"),
    ("「夢リスト」を書き始めたら人生が変わったのか検証", "カナエル直結", "夢リストの効果を自分で検証。カナエルへの自然な誘導。"),
    ("夢と目標の違い（FIREして気づいたこと）", "カナエル直結", "FIREを達成して初めてわかった「夢」と「目標」の本質的な差。"),
]

start_row = ws.max_row + 1

for i, (title, category, memo) in enumerate(new_topics):
    row_num = start_row + i
    values = [row_num - 1, "", title, memo, "予定", ""]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
        cell.fill = planned_fill
    # Status cell
    ws.cell(row=row_num, column=5).font = Font(name="Arial", size=10, bold=True)
    ws.cell(row=row_num, column=5).alignment = Alignment(horizontal="center", vertical="top")
    # Category in column 7 (extra column)
    cat_cell = ws.cell(row=row_num, column=7, value=category)
    cat_cell.font = Font(name="Arial", size=10, color="666666")
    cat_cell.alignment = Alignment(vertical="top")
    cat_cell.border = border
    cat_cell.fill = planned_fill
    ws.row_dimensions[row_num].height = 45

# Add header for category column if not present
if ws.cell(row=1, column=7).value is None:
    from openpyxl.styles import Font as F, PatternFill as PF, Alignment as Al
    h = ws.cell(row=1, column=7, value="カテゴリ")
    h.font = F(name="Arial", bold=True, color="FFFFFF", size=11)
    h.fill = PF("solid", start_color="1F3864")
    h.alignment = Al(horizontal="center", vertical="center")
    h.border = border
    ws.column_dimensions["G"].width = 18

wb.save("/Users/katsuo/my-projects/sns-strategy/planning/週刊かつお_ネタ帳.xlsx")
print("Done:", len(new_topics), "件追加")
