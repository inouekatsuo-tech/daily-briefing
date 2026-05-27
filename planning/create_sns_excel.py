from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta

wb = Workbook()

# ==================== Helper styles ====================
def header_font(color="FFFFFF"):
    return Font(name="Arial", bold=True, color=color)

def regular_font():
    return Font(name="Arial")

def center_align():
    return Alignment(horizontal="center", vertical="center")

def right_align():
    return Alignment(horizontal="right", vertical="center")

def left_align():
    return Alignment(horizontal="left", vertical="center")

def make_fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = max(width, 15)

def freeze_first_row(ws):
    ws.freeze_panes = "A2"

def style_header_row(ws, row, num_cols, fill_color, font_color="FFFFFF"):
    fill = make_fill(fill_color)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name="Arial", bold=True, color=font_color)
        cell.fill = fill
        cell.alignment = center_align()

# ==================== Sheet 1: フォロワー推移 ====================
ws1 = wb.active
ws1.title = "📊 フォロワー推移"
ws1.sheet_properties.tabColor = "4472C4"  # Blue

headers1 = [
    "日付（週次）", "X フォロワー数", "X 前週比",
    "Instagram フォロワー数", "Instagram 前週比",
    "Threads フォロワー数", "Threads 前週比",
    "note フォロワー数", "note 前週比",
    "合計フォロワー数", "合計前週比"
]
for col, h in enumerate(headers1, 1):
    ws1.cell(row=1, column=col, value=h)
style_header_row(ws1, 1, len(headers1), "808080")

# Sample data: 5 weeks starting 2026/03/25
base_date = date(2026, 3, 25)
x_start, ig_start, th_start, note_start = 150, 320, 80, 1494
weekly_x = [0, 12, 15, 10, 18]
weekly_ig = [0, 22, 18, 25, 20]
weekly_th = [0, 8, 10, 6, 12]
weekly_note = [0, 15, 20, 12, 18]

rows_data = []
for i in range(5):
    d = base_date + timedelta(weeks=i)
    xf = x_start + sum(weekly_x[:i+1])
    ig = ig_start + sum(weekly_ig[:i+1])
    th = th_start + sum(weekly_th[:i+1])
    nt = note_start + sum(weekly_note[:i+1])
    rows_data.append((d, xf, ig, th, nt))

for i, (d, xf, ig, th, nt) in enumerate(rows_data):
    r = i + 2
    ws1.cell(row=r, column=1, value=d).number_format = "yyyy/mm/dd"
    ws1.cell(row=r, column=2, value=xf)
    if r == 2:
        ws1.cell(row=r, column=3, value="-")
    else:
        cell = ws1.cell(row=r, column=3)
        cell.value = f"=B{r}-B{r-1}"
        cell.number_format = "+0;-0;0"
    ws1.cell(row=r, column=4, value=ig)
    if r == 2:
        ws1.cell(row=r, column=5, value="-")
    else:
        cell = ws1.cell(row=r, column=5)
        cell.value = f"=D{r}-D{r-1}"
        cell.number_format = "+0;-0;0"
    ws1.cell(row=r, column=6, value=th)
    if r == 2:
        ws1.cell(row=r, column=7, value="-")
    else:
        cell = ws1.cell(row=r, column=7)
        cell.value = f"=F{r}-F{r-1}"
        cell.number_format = "+0;-0;0"
    ws1.cell(row=r, column=8, value=nt)
    if r == 2:
        ws1.cell(row=r, column=9, value="-")
    else:
        cell = ws1.cell(row=r, column=9)
        cell.value = f"=H{r}-H{r-1}"
        cell.number_format = "+0;-0;0"
    total_cell = ws1.cell(row=r, column=10)
    total_cell.value = f"=B{r}+D{r}+F{r}+H{r}"
    if r == 2:
        ws1.cell(row=r, column=11, value="-")
    else:
        cell = ws1.cell(row=r, column=11)
        cell.value = f"=J{r}-J{r-1}"
        cell.number_format = "+0;-0;0"

# Style data rows
for r in range(2, 7):
    ws1.cell(row=r, column=1).font = regular_font()
    ws1.cell(row=r, column=1).alignment = center_align()
    for col in [2, 3, 4, 5, 6, 7, 8, 9, 10, 11]:
        cell = ws1.cell(row=r, column=col)
        cell.font = regular_font()
        cell.alignment = right_align()

col_widths1 = [18, 18, 12, 22, 18, 20, 18, 18, 14, 18, 14]
for col, w in enumerate(col_widths1, 1):
    set_col_width(ws1, col, w)

freeze_first_row(ws1)

# ==================== Sheet 2: 投稿カレンダー ====================
ws2 = wb.create_sheet("📅 投稿カレンダー")
ws2.sheet_properties.tabColor = "70AD47"  # Green

headers2 = [
    "日付", "曜日", "X投稿内容（予定）", "X投稿タイプ",
    "Instagram投稿内容（予定）", "Instagramタイプ",
    "Threads投稿内容（予定）", "note記事タイトル（予定）",
    "投稿済みチェック（✓）", "メモ"
]
for col, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=col, value=h)
style_header_row(ws2, 1, len(headers2), "1F3864")

weekday_ja = ["月", "火", "水", "木", "金", "土", "日"]
start = date(2026, 3, 25)
for i in range(28):
    d = start + timedelta(days=i)
    r = i + 2
    cell_d = ws2.cell(row=r, column=1, value=d)
    cell_d.number_format = "yyyy/mm/dd"
    cell_d.font = regular_font()
    cell_d.alignment = center_align()
    wd = ws2.cell(row=r, column=2, value=weekday_ja[d.weekday()])
    wd.font = regular_font()
    wd.alignment = center_align()
    for col in range(3, 11):
        ws2.cell(row=r, column=col).font = regular_font()
        ws2.cell(row=r, column=col).alignment = left_align()

col_widths2 = [15, 8, 30, 20, 30, 20, 30, 30, 18, 25]
for col, w in enumerate(col_widths2, 1):
    set_col_width(ws2, col, w)

freeze_first_row(ws2)

# ==================== Sheet 3: ネタストック ====================
ws3 = wb.create_sheet("💡 ネタストック")
ws3.sheet_properties.tabColor = "FFD700"  # Yellow

headers3 = ["No.", "カテゴリ", "タイトル・テーマ", "内容メモ", "適したSNS", "優先度", "ステータス", "使用日"]
for col, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=col, value=h)
style_header_row(ws3, 1, len(headers3), "7F6000", "FFFFFF")

ideas = [
    (1, "バツ2", "バツ2からわかった「結婚に向かない人」の特徴", "自分の失敗談を元に、結婚に向いていないタイプを分析。共感狙い。", "X/Threads", "高", "未使用", ""),
    (2, "バツ2", "離婚経験者が語る「やり直せるなら変えること3つ」", "後悔していることと今の考え方の変化。エピソード形式。", "note/X", "高", "未使用", ""),
    (3, "世界一周", "チリで出会った人生を変えた一言", "旅先での出会いと気づき。感動系エピソード。", "X/Instagram/note", "高", "未使用", ""),
    (4, "世界一周", "世界一周してわかった「日本人が損してること」", "海外と日本の文化・思考の違いを比較。逆張り系。", "X/Threads", "高", "未使用", ""),
    (5, "世界一周", "一人旅で孤独と戦った夜の話", "旅の孤独感とそれを乗り越えた体験。リアルなエピソード。", "Instagram/note", "中", "未使用", ""),
    (6, "起業", "ゼロから起業した最初の3ヶ月で犯した失敗", "初期の失敗談と学び。知識系+エピソード系。", "note/X", "高", "未使用", ""),
    (7, "FIRE", "FIREを目指して気づいた「お金より大事なもの」", "FIRE達成に向かう中での価値観の変化。人生論系。", "X/Threads/note", "高", "未使用", ""),
    (8, "FIRE", "サラリーマンを辞めたら収入が3倍になった話", "起業・副業でのリアルな収入変化。逆張り+本音系。", "X", "高", "未使用", ""),
    (9, "カナエル", "「人生を変えたい」と思ったときに最初にすべきこと", "カナエルの価値観をベースにした行動論。", "X/Threads/note", "高", "未使用", ""),
    (10, "SNS", "フォロワー0から始めた私がやって良かったこと悪かったこと", "SNS運用の試行錯誤。知識+本音系。", "X/Threads", "中", "未使用", ""),
    (11, "人生論", "40代で人生をリセットして気づいた「当たり前の嘘」", "社会の常識への疑問と自分なりの答え。逆張り系。", "X/Threads/note", "中", "未使用", ""),
    (12, "バツ2", "再婚を選ばなかった理由と今の幸せの形", "個人的な選択と価値観。共感・本音系。", "note/Instagram", "中", "未使用", ""),
]

for idea in ideas:
    r = idea[0] + 1
    for col, val in enumerate(idea, 1):
        cell = ws3.cell(row=r, column=col, value=val)
        cell.font = regular_font()
        if col in [1, 6, 7]:
            cell.alignment = center_align()
        else:
            cell.alignment = left_align()

col_widths3 = [8, 20, 35, 45, 25, 10, 12, 15]
for col, w in enumerate(col_widths3, 1):
    set_col_width(ws3, col, w)

freeze_first_row(ws3)

# ==================== Sheet 4: 投稿パフォーマンス ====================
ws4 = wb.create_sheet("📈 投稿パフォーマンス")
ws4.sheet_properties.tabColor = "ED7D31"  # Orange

headers4 = [
    "日付", "SNS", "投稿内容（冒頭）", "いいね数",
    "リポスト/シェア数", "コメント数", "インプレッション数",
    "エンゲージメント率", "フォロワー増減", "バズったか", "学び・メモ"
]
for col, h in enumerate(headers4, 1):
    ws4.cell(row=1, column=col, value=h)
style_header_row(ws4, 1, len(headers4), "375623")

# Leave rows 2+ empty for user to fill in, but set format for engagement rate
for r in range(2, 52):
    cell = ws4.cell(row=r, column=8)
    cell.value = f"=IFERROR((D{r}+E{r}+F{r})/G{r},\"\")"
    cell.number_format = "0.0%"
    cell.font = regular_font()
    for col in [1, 2, 3, 4, 5, 6, 7, 9, 10, 11]:
        ws4.cell(row=r, column=col).font = regular_font()

col_widths4 = [15, 12, 35, 10, 18, 12, 20, 18, 15, 12, 35]
for col, w in enumerate(col_widths4, 1):
    set_col_width(ws4, col, w)

freeze_first_row(ws4)

# ==================== Sheet 5: 月次目標 ====================
ws5 = wb.create_sheet("🎯 月次目標")
ws5.sheet_properties.tabColor = "FF0000"  # Red

headers5 = [
    "月",
    "Xフォロワー目標", "X実績", "X達成率",
    "Instagramフォロワー目標", "Instagram実績", "Instagram達成率",
    "Threadsフォロワー目標", "Threads実績", "Threads達成率",
    "noteフォロワー目標", "note実績", "note達成率",
    "合計目標", "合計実績", "合計達成率",
    "月次コメント"
]
for col, h in enumerate(headers5, 1):
    ws5.cell(row=1, column=col, value=h)
style_header_row(ws5, 1, len(headers5), "C00000")

# Monthly targets 2026/04 - 2026/12
x_base = 150
ig_base = 320
th_base = 80
note_base = 1494

months = [
    ("2026年4月", 250, 470, 160, 1594),
    ("2026年5月", 350, 620, 240, 1694),
    ("2026年6月", 450, 770, 320, 1794),
    ("2026年7月", 550, 920, 400, 1894),
    ("2026年8月", 650, 1070, 480, 1994),
    ("2026年9月", 750, 1220, 560, 2094),
    ("2026年10月", 850, 1370, 640, 2194),
    ("2026年11月", 950, 1520, 720, 2294),
    ("2026年12月", 1050, 1670, 800, 2394),
]

for i, (month, x_tgt, ig_tgt, th_tgt, note_tgt) in enumerate(months):
    r = i + 2
    ws5.cell(row=r, column=1, value=month).font = regular_font()
    ws5.cell(row=r, column=1).alignment = center_align()

    ws5.cell(row=r, column=2, value=x_tgt).font = regular_font()
    ws5.cell(row=r, column=2).alignment = right_align()
    ws5.cell(row=r, column=3).font = regular_font()  # 実績は空白
    ws5.cell(row=r, column=3).alignment = right_align()
    cell = ws5.cell(row=r, column=4)
    cell.value = f"=IFERROR(C{r}/B{r},\"\")"
    cell.number_format = "0.0%"
    cell.font = regular_font()
    cell.alignment = right_align()

    ws5.cell(row=r, column=5, value=ig_tgt).font = regular_font()
    ws5.cell(row=r, column=5).alignment = right_align()
    ws5.cell(row=r, column=6).font = regular_font()
    ws5.cell(row=r, column=6).alignment = right_align()
    cell = ws5.cell(row=r, column=7)
    cell.value = f"=IFERROR(F{r}/E{r},\"\")"
    cell.number_format = "0.0%"
    cell.font = regular_font()
    cell.alignment = right_align()

    ws5.cell(row=r, column=8, value=th_tgt).font = regular_font()
    ws5.cell(row=r, column=8).alignment = right_align()
    ws5.cell(row=r, column=9).font = regular_font()
    ws5.cell(row=r, column=9).alignment = right_align()
    cell = ws5.cell(row=r, column=10)
    cell.value = f"=IFERROR(I{r}/H{r},\"\")"
    cell.number_format = "0.0%"
    cell.font = regular_font()
    cell.alignment = right_align()

    ws5.cell(row=r, column=11, value=note_tgt).font = regular_font()
    ws5.cell(row=r, column=11).alignment = right_align()
    ws5.cell(row=r, column=12).font = regular_font()
    ws5.cell(row=r, column=12).alignment = right_align()
    cell = ws5.cell(row=r, column=13)
    cell.value = f"=IFERROR(L{r}/K{r},\"\")"
    cell.number_format = "0.0%"
    cell.font = regular_font()
    cell.alignment = right_align()

    cell = ws5.cell(row=r, column=14)
    cell.value = f"=B{r}+E{r}+H{r}+K{r}"
    cell.font = regular_font()
    cell.alignment = right_align()

    cell = ws5.cell(row=r, column=15)
    cell.value = f"=IFERROR(C{r}+F{r}+I{r}+L{r},\"\")"
    cell.font = regular_font()
    cell.alignment = right_align()

    cell = ws5.cell(row=r, column=16)
    cell.value = f"=IFERROR(O{r}/N{r},\"\")"
    cell.number_format = "0.0%"
    cell.font = regular_font()
    cell.alignment = right_align()

    ws5.cell(row=r, column=17).font = regular_font()
    ws5.cell(row=r, column=17).alignment = left_align()

col_widths5 = [15, 20, 12, 12, 25, 18, 22, 22, 18, 22, 20, 12, 16, 12, 12, 16, 25]
for col, w in enumerate(col_widths5, 1):
    set_col_width(ws5, col, w)

freeze_first_row(ws5)

# Save
output_path = "/Users/katsuo/my-projects/sns-strategy/planning/SNS戦略管理.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
